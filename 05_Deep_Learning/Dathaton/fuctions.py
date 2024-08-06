import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

class functions:
    @staticmethod
    def filter_columns(df, filters: list): 
        selected_columns = [True] * len(df.columns)  # Inicializa todas as colunas como True
        for index, column in enumerate(df.columns):
            if any(filter in column for filter in filters):
                selected_columns[index] = False
        return df[df.columns[selected_columns]]

    @staticmethod
    def cleaning_dataset(df):
        _df = df.dropna(subset=df.columns.difference(['NOME']), how='all') # executa o dropna para todas as colunas sem visualizar a coluna NOME
        _df = _df[~_df.isna().all(axis=1)] # remove linhas com apenas NaN, se tiver algum dado na linha não remove
        return _df

    @staticmethod
    def plot_exact_counter(size, x, y, df) -> None:
        plt.figure(figsize=size)
        barplot = plt.bar(y.index, y.values)
        plt.xlabel(x)
        plt.ylabel('Count')

        for index, value in enumerate(y.values):
            plt.text(index, value, round(value, 2), color='black', ha="center")

        plt.show()

    @staticmethod
    def dummie_int(df, list_columns):
        _df = df.copy()  # Adiciona cópia para evitar a modificação direta do DataFrame original
        for column in list_columns:
            _df[column] = _df[column].replace(['Sim', 'Não'], [1,0])
        return _df

    @staticmethod
    # Função para carregar o conteúdo HTML
    def load_html(file_path):
        with open(file_path, 'r', encoding='utf-8') as file:
            return file.read()

    @staticmethod
    # Função para carregar o conteúdo CSS
    def load_css(file_path):
        with open(file_path, 'r', encoding='utf-8') as file:
            return file.read() 
    
    @staticmethod
    def create_sample_workbook():
        # Cria um workbook e uma planilha
        wb = Workbook()
        ws = wb.active
        ws.title = "Modelo"

        # Define as colunas
        columns = [
            'IAA_2022', 'IEG_2022', 'IPS_2022', 'IDA_2022', 'NOTA_PORT_2022',
            'NOTA_MAT_2022', 'NOTA_ING_2022', 'IPP_2022', 'INDICADO_BOLSA_2022'
        ]

        # Adiciona os cabeçalhos na planilha
        for col_num, column_title in enumerate(columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)

        # Salva o workbook em um objeto BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output
